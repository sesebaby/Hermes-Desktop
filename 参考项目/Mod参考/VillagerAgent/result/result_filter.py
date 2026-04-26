import os
import json
from collections import defaultdict
import openpyxl
import re
from openpyxl.styles import Font, Alignment
import shutil

# 定义路径
# base_dir = './'
base_dir = './../processed_result/base_agent_multi/qwen3-80b-instruct/supplement/1/'
output_excel_path = 'SFT base agent.xlsx'

def base_agent_filter():

    # 初始化统计数据
    score_100_distribution = defaultdict(int)  # score 为 100 的文件夹数量

    # 遍历当前路径下的所有文件夹
    for folder_name in os.listdir(base_dir):
        if os.path.isdir(os.path.join(base_dir, folder_name)):
            # 解析文件夹名称
            parts = folder_name.split('_')
            category = parts[0]

            # 如果是 interact 类别，进一步解析子类别
            if category == 'interact' and len(parts) > 1:
                subcategory = parts[1]
                category = f"interact_{subcategory}"  # 使用 interact_子类别 作为类别名

            # 检查 score.json 文件
            score_json_path = os.path.join(base_dir, folder_name, 'score.json')
            folder_path = os.path.join(base_dir, folder_name)
            if os.path.exists(score_json_path):
                with open(score_json_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                # 检查是否有 score 字段且值为 100
                if data.get('score') == 100:
                    score_100_distribution[category] += 1
                # else:
                #     # 删除不为100的文件夹
                #     shutil.rmtree(folder_path)  # <-- 修改：删除该文件夹
                #     print(f"已删除文件夹：{folder_name}")  # <-- 修改：日志提示

    # 创建 Excel 文件
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QWEN"

    # 写入表头
    ws.append(["类别", "score 为 100 的文件夹数量"])

    # 写入 score 为 100 的分布情况
    for category, count in score_100_distribution.items():
        ws.append([category, count])

    # 保存 Excel 文件
    wb.save(output_excel_path)

    print(f"统计结果已保存到 {output_excel_path}")


def base_agent_multi_filter():
    folders = [os.path.join(base_dir, f) for f in os.listdir(base_dir) if os.path.isdir(os.path.join(base_dir, f))]
    construction_results = []
    farming_results = []

    for folder in sorted(folders):  # 按字典序
        score_path = os.path.join(folder, "score.json")
        if not os.path.exists(score_path):
            continue

        match = re.search(r"task(\d+)", folder)
        task_idx = int(match.group(1)) if match else None

        with open(score_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        if "construction" in folder.lower():
            construction_results.append({
                "task_idx": task_idx,
                "block_hit_rate": data.get("block_hit_rate", 0.0),
                "view_hit_rate": data.get("view_hit_rate", 0.0)
            })
        elif "farming" in folder.lower():
            farming_results.append({
                "task_idx": task_idx,
                "score": data.get("score", 0.0),
                "cooperation": data.get("cooperation", 0.0),
                "efficiency": data.get("efficiency", 0.0),
                "balance": data.get("balance", 0.0)
            })

    wb = openpyxl.Workbook()

    # 样式定义
    font = Font(name="SimHei")  # 黑体
    align = Alignment(horizontal="center", vertical="center")

    # ===== Construction Sheet =====
    ws1 = wb.active
    ws1.title = "construction"
    ws1.append(["task_idx", "bhr", "vhr"])
    for r in sorted(construction_results, key=lambda x: x["task_idx"]):
        ws1.append([
            r["task_idx"],
            round(r["block_hit_rate"], 3),
            round(r["view_hit_rate"], 3)
        ])
    if construction_results:
        avg_block = sum(r["block_hit_rate"] for r in construction_results) / len(construction_results)
        avg_view = sum(r["view_hit_rate"] for r in construction_results) / len(construction_results)
        ws1.append(["Average", round(avg_block, 3), round(avg_view, 3)])

    # ===== Farming Sheet =====
    ws2 = wb.create_sheet("farming")
    ws2.append(["task_idx", "score", "cooperation", "efficiency", "balance"])
    for r in sorted(farming_results, key=lambda x: x["task_idx"]):
        ws2.append([r["task_idx"], round(r["score"], 3), round(r["cooperation"], 3), round(r["efficiency"], 3), round(r["balance"], 3)])
    if farming_results:
        avg_score = sum(r["score"] for r in farming_results) / len(farming_results)
        avg_cooperation = sum(r["cooperation"] for r in farming_results) / len(farming_results)
        avg_efficiency = sum(r["efficiency"] for r in farming_results) / len(farming_results)
        avg_balance = sum(r["balance"] for r in farming_results) / len(farming_results)
        ws2.append(["Average", round(avg_score, 3), round(avg_cooperation, 3), round(avg_efficiency, 3), round(avg_balance, 3)])

    # ===== 应用格式 =====
    for ws in [ws1, ws2]:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
                cell.alignment = align

    wb.save(base_dir + "base agent multi.xlsx")
    print("✅ 结果已保存到 base agent multi.xlsx")

# 运行
if __name__ == "__main__":
    # base_agent_filter()
    base_agent_multi_filter()
